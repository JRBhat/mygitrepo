from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook


from Internal_Imports_Stable import studynumber, excelfile, Test_type
from Internal_Imports_Stable import header, header_draft, pagestyle, hypersetup
from Latex_File_Create_Stable import create_final_latex_file
# from Common_Functions import create_Latex_document
# from Common_Functions_Refactored import create_Latex_document
from Common_Functions_Stable import create_Latex_document


def transpositioned(area_code_list, time_code_list, area_sorted, time_sorted, sub_sorted, 
                                filepaths, filenamelist, rownames, colnames, transfer_list, draft_flag, random_iter,random=True):
    """
    Transpose of standard template Rows and columns are swapped in excel workseet
    """    

    wb = load_workbook(excelfile)

    ws4 = wb['Modify_Template_here']

    nr_col= len(area_code_list)#len(time_list)
    nr_row= len(time_code_list)#len(area_list)

    max_height=round(1.0/(1.25*nr_row), 2)
    max_width=round(1.0/(1.5*nr_col), 2)

# delete
    transfer_list_transposed = []
    for time_code in time_sorted:
        for area_code in area_sorted:
            for transfer_value in transfer_list:
                if area_code in transfer_value and time_code in transfer_value:
                    transfer_list_transposed.append(transfer_value)

    Transpose_Iter = iter(transfer_list_transposed) # initialize iterator

    for ttn, _ in enumerate(time_code_list, start=2):   # Replacing the item in rows/columns  in modified excel sheet with the derandomised value for the first subject
        for aan, _ in enumerate(area_code_list, start=2): # TODO: For custom template - value in the rows and col must be read first and then rand value must be swapped with it
            ws4.cell(row=ttn, column=aan).value = next(Transpose_Iter)

    wb.save(excelfile)

    if random == True:
        
        lax_document, new_page_alert = create_Latex_document(colnames, rownames, sub_sorted, max_height, max_width, 
                                                                time_sorted, area_sorted, filepaths, filenamelist, 
                                                                Type="Standard", RandomList=random_iter)
    else:
        lax_document, new_page_alert = create_Latex_document(colnames, rownames, sub_sorted, max_height, max_width, 
                                                                time_sorted, area_sorted, filepaths, filenamelist, 
                                                                Type="Standard", RandomList=random_iter)
    
    transpose_tex_file = create_final_latex_file(studynumber, header, header_draft, pagestyle, 
                                                 hypersetup, lax_document, new_page_alert, colnames, 
                                                 Test_type, draft=draft_flag)#, special=special)
    
    
    return transpose_tex_file
