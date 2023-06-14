import os
import shutil
import datetime
import subprocess
import glob 
from time import sleep
import sys

from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook, utils
from ImageAnalysis import Util
from Internal_Imports_Stable import validated_path, randomfilepath, Beirsdorf_randomfilepath, filename_mask, excelfile
from Internal_Imports_Stable import  studynumber, file_extension, draft_flag, dummy_log_file, Test_number, Test_type
from Excel_Create_refactoring import create_excel 
from Template_Standard_Stable import standardize
from Template_Custom_Stable import customize
from Template_Transpose_Stable import transpositioned
from Randomization_Template_Stable import randomize_std
from Randomization_Template_Transposed_Stable import randomize_transp
from Beirsdorf_Template_Stable import do_the_Beirsdorf_transform
from Column_Name_Per_Row_Template import give_each_row_columnames
from Common_Functions_Stable import get_row_and_columns, replace_missing_barcodes_with_dummy, create_code_dict_for_excel_table
from Insert_Description_Stable import Insert_description_file
import re

from Helper_modules.file_renamers.Visia_file_rename_Utility import rename_visia_files


def main():
    """
    the heart of the program containing bulk of the logic 
    """
    
    filenamelist = Util.getAllFiles(validated_path, file_extension, depth=-1)

    # ............... check for visia files.........
    Visia_flag = 0
    fn_regx = re.compile(rf'{filename_mask}')
    visia_rgx_mask = re.compile(r"([0-9]*)_(T[0-9]{2})_([a-zA-Z \-]*)_([a-zA-Z1-2 \-]*)")

    first_img_filename = [fn for fn in filenamelist if fn.endswith(".jpg")][0].split("\\")[-1]
        
    if not re.match(fn_regx, first_img_filename):
        if visia_rgx_mask.search(first_img_filename).group(0):
            Test_type = f"TestCustom_{Test_number}"
            Visia_flag = 1
            rename_visia_files(validated_path, studynumber) 
        else:
            print("Regex match not matching for visia or std mask; check filename again")
            sys.exit(1)

    # .............Missing file handle.......................................

    filenamelist = Util.getAllFiles(validated_path, file_extension, depth=-1)
    dummy_list = replace_missing_barcodes_with_dummy(filenamelist, Visia_flag) 

    if dummy_list != None: # if missing files found
        filenamelist = Util.getAllFiles(validated_path, file_extension, depth=-1) # after handling missing data

        # logging cloned files to handle missing barcodes
        
        with open(os.path.join(validated_path, dummy_log_file), "w") as log_f:
            for item in dummy_list:
                log_f.write(item + "\n")

    #................... Excel layout template creation..........................
    
    dict_code = create_code_dict_for_excel_table(validated_path)
    
    area_code_list, time_code_list, area_sorted, \
    time_sorted, sub_sorted, filepaths, filenamelist, \
    Main_mapping_dict, transfer_list = create_excel(validated_path, excelfile, dict_code)


    # console output - excel process confirmation
    open_path = os.getcwd()
    proc_excel = subprocess.Popen(f"start {os.path.join(open_path, excelfile)}", shell=True)
    proc_excel.wait()
    input("folder open..Press enter")

    # Confirms user input
    while True:
        Check_user_input = input("Have you entered all necessary study information(Y/N)?: ")
        if Check_user_input == 'Y' or Check_user_input == 'y': 
            print(f"Reading from Modified template worksheet of the excel file {excelfile}")
            sleep(1)
            break
        
        elif Check_user_input == 'N' or Check_user_input == 'n':
            print("Stopping program")
            exit(1)


# ........................Reading from Excel file after user has made changes..................
    # after user says 'y'
    wb = load_workbook(excelfile)
    # opens modified worksheet
    ws4 = wb['Modify_Template_here']
    # reading cell A1 - only for CustomTemplate.py
    cell_a1 = ws4.cell(row=1, column=1).value  # checking if a * or § is found in this cell
    print(cell_a1)

    # condition handling column names for each row
    column_row_name_list = []
    for coln, col in enumerate(ws4.columns):
        for cell in col:
            if cell.value == "§": # checking for this flag - option to include name above each image
                print("found")
                special_col_index = utils.get_column_letter(cell.column)

                column_row_name_list = []
                for row in ws4[special_col_index]:
                    if row.value == "None" or row.value == None or row.value == "":
                        break
                    column_row_name_list.append(row.value)
                column_row_name_list = column_row_name_list[1:]
                print(column_row_name_list)
                break
        if coln == 20:
            break

    if cell_a1 == "*": # custom template flag = *

        randomized_files_iterator = None
        if randomfilepath is not None:
            try: 
                randomized_files_iterator = randomize_std(filepaths, randomfilepath, area_sorted, time_sorted, sub_sorted)
            except IndexError:
                print("Custom - file List index out of range; Check the random file again for duplicates or Missing values")
                exit(1)
        rownames, colnames = get_row_and_columns(ws4, 3, 2)
        tex_file = customize(area_sorted, time_sorted, sub_sorted, rownames, colnames, 
            Main_mapping_dict, draft_flag, randomized_files_iterator)
    
    else:
        rownames, colnames = get_row_and_columns(ws4, 2, 1)
        if time_sorted[0] in rownames[0]:
            if randomfilepath is not None:
                try: 
                    randomized_files_iterator, filepathrandlist = randomize_transp(filepaths, randomfilepath, area_sorted, time_sorted, sub_sorted)
                    
                    path_sep_list = [path[0] for path in filepathrandlist] # retrievs paths from the tuple
                    name_sep_file_list = [str_path.split("\\")[-1] for str_path in path_sep_list] # retrieves names from paths
                    tex_file = transpositioned(area_code_list, time_code_list, area_sorted, time_sorted, sub_sorted, 
                                                path_sep_list, name_sep_file_list, rownames, colnames, transfer_list, draft_flag, randomized_files_iterator)
                except IndexError:
                    print("List index out of range; Check the random file again for duplicates or Missing values")
                    exit(1)
            else:
                randomized_files_iterator = None
                tex_file = transpositioned(area_code_list, time_code_list, area_sorted, time_sorted, sub_sorted, 
                                            filepaths, filenamelist, rownames, colnames, transfer_list, draft_flag, randomized_files_iterator)
        
        else:
            if randomfilepath is not None:
                try: 
                    randomized_files_iterator = randomize_std(filepaths, randomfilepath, area_sorted, time_sorted, sub_sorted)
                except IndexError:
                    print("List index out of range; Check the random file again for duplicates or Missing values")
                    exit(1)
            else:
                randomized_files_iterator = None
            tex_file = standardize(area_code_list, time_code_list, area_sorted, time_sorted, sub_sorted, 
                                    filepaths, filenamelist, rownames, colnames, draft_flag, randomized_files_iterator)

    ## Additional features/study specific layouts
    if len(column_row_name_list) > 0:
        tex_file = give_each_row_columnames(column_row_name_list, tex_file)

    appended_beiers_fpath = None
    if Beirsdorf_randomfilepath != None:
        _, appended_beiers_fpath = do_the_Beirsdorf_transform(tex_file, Beirsdorf_randomfilepath)

    # A function for inserting the description page
    if appended_beiers_fpath != None:
        final_tex_file = Insert_description_file(appended_beiers_fpath.split("\\")[-1])
    else:
        final_tex_file = Insert_description_file(tex_file)

    # Generate the PDF from tex file automatically and display it to the user
    if os.path.isfile(final_tex_file):
        try:
            proc1 = subprocess.Popen(f"pdflatex -interaction=nonstopmode -halt-on-error {final_tex_file}  && pause", shell=True)
            proc1.wait()
            print("All files generated...starting archiving and cleaning process")
            sleep(1)
        except:
            os.startfile(final_tex_file)
            input("Press enter to exit...")

# ....................................................................................................
# ............................THE END........THE END..........THE END........THE END.......................................
# .........................................................................................................................
if __name__ == '__main__':

    main()
    # deleting unneccessary files
    for f in os.listdir(os.getcwd()):
        if f.endswith("gz") or f.endswith("log") or f.endswith("aux") or f.endswith("out"):
            os.remove(os.path.join(os.getcwd(), f))
            print(f"{f} deleted")

    test_folder_path = os.path.join(validated_path, "Test_results")
    if not os.path.isdir(test_folder_path):
        os.mkdir(test_folder_path)

    move_list = []
    for f in os.listdir(os.getcwd()):
        descp_folder = r"D:\Code\Software_test_sample_data\Dev_Proj_5__ExcelToPdfConverter\DESCPR_files"
        if f.endswith("docx") or (f.endswith("tex") and not "with" in f and "description" in f):
            shutil.move(os.path.join(os.getcwd(), f), os.path.join(descp_folder,f))
        elif f.endswith("pdf") or f.endswith("tex") or f.endswith("xlsx") or (f.endswith(".txt") and "missing" in f):
            move_list.append(f)

    # archiving all files into the study folder "Test_results" within the images folder
    if len(move_list) > 0:
        while True:
            archive_approval = input("Would you like to archive the files for this test(y/n): ")     
            if archive_approval == "y":
                timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                new_folderpath = os.path.join(test_folder_path, f"{studynumber}_{Test_type}_{file_extension[2:]}__{timestamp}")
                os.mkdir(new_folderpath)
                for selected_file in move_list:
                    shutil.move(os.path.join(os.getcwd(), selected_file), os.path.join(new_folderpath, selected_file)) 
                    print(f"{selected_file} moved to subfolder - {studynumber}_{Test_type}_{file_extension[2:]}__{timestamp} inside folder - Test_results")
                subprocess.Popen(f"start {new_folderpath} && pause", shell=True)
                type_pdf = "\*pdf"
                type_latex = "\*tex"
                pdf_files = glob.glob(new_folderpath + type_pdf)
                latex_files = glob.glob(new_folderpath + type_latex)
                try: 
                    max_file_pdf = max(pdf_files, key=os.path.getctime)
                    max_file_latex = max(latex_files, key=os.path.getctime)

                    if max_file_latex != None or max_file_pdf != None:
                        if max_file_latex != None and max_file_pdf == None:
                            proc2 = subprocess.Popen(f"d: && start {max_file_latex} && pause", shell=True)
                            proc2.wait()
                            print("Goodbye")
                        proc2 = subprocess.Popen(f"d: && start {max_file_latex} && start {max_file_pdf} && pause", shell=True)
                        proc2.wait()
                        print("Goodbye")

                except ValueError:
                    print("No pdf or latex file found")
                    print("Goodbye")
                break
    sys.exit(0)