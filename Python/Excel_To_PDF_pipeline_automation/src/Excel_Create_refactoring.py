
from openpyxl import Workbook
import os
import re
from ImageAnalysis import Util
import sys
from Internal_Imports_Stable import  file_extension, filename_mask


def create_excel(mypath, excelfile, dict_code):
    """
    Creates Excel file where User can adjust layout and enter additional names and inputs
    """

    subj = re.compile(r'S[0-9]{3}')
    area = re.compile(r'F[0-9]{2}')
    time = re.compile(r'T[0-9]{2}')

    file_key_mask = re.compile(rf'{filename_mask}')

    # list of filepaths
    filenamelist = Util.getAllFiles(mypath, file_extension, depth=-1) 
    
    ######################  SHEET 1 : Raw data storage ############################################
    # list of filepaths
    
    # New excel file
    wb = Workbook()
    ws1= wb['Sheet']
    ws1.title = 'Raw_Data' 

    # print filenames on each row in excel sheet 1
    for row in filenamelist:
        ws1.append((row, os.path.split(row)[-1]))
    wb.save(excelfile)

    ######################  SHEET 2 : Stores full template  #######################################
    ws2 = wb.create_sheet('Template_data', 2)
    WholePathList = [fpath for fpath in filenamelist]
    Wholetargetpathlist = []
    dictTargetPathlist = []

    # get only last three path elements
    for Pth in WholePathList:
        temp_list = Pth.split('\\')#use PathLib
        Wholetargetpathlist.append('\\'.join(temp_list[-3:]))
    for ppth in Wholetargetpathlist:
        path_part = '\\'.join(ppth.split('\\')[:-1])
        try: 
            filename_part = file_key_mask.search(ppth).group(0)
        except AttributeError:
            print("The provided Filemask does not retrieve the given filename. \
                  Please adjust the REGEX mask in the json file according to the image names.")
            sys.exit(1)
        dictTargetPathlist.append('\\'.join([path_part, filename_part]))
    print(dictTargetPathlist)
    # print(f'target paths: {targetpathlist}')

    ## Extraction and group assignment
    # Regex masks for subject, area and time codes from filenames# 

    subject_group = []
    area_group = []
    time_group = []
    # read subj ids from filenames, then group & sort
    for filpth in filenamelist:
        subjtemplist = filpth.split('\\')
        # print(subjtemplist[1])
        subject_group.append(subj.search(subjtemplist[-1]).group(0))

    sub_sorted = sorted(set(subject_group))
    # print(sorted(set(subject_group)))

    # Tuple attaches subject number to transfer list element
    key_list = []
    for i, subno in enumerate(subject_group):
        # key_list.append((Wholetargetpathlist[i][:-4], subno))
        key_list.append((dictTargetPathlist[i], subno))

    # dictionary maps key_list with filenames
    Main_mapping_dict = {}
    for no, key in enumerate(key_list):
        Main_mapping_dict[key] = WholePathList[no]

    for row in zip(WholePathList, Wholetargetpathlist, dictTargetPathlist):
        ws2.append(row)
    wb.save(excelfile)

    ######################  SHEET 2b  #########################################################
    # links path elements with Template (extracts subject 1 elements only)
    ws2b = wb.create_sheet('First_Subject_tempData', 3)
    sub1Pathlist = [fpath for fpath in filenamelist if sub_sorted[0] in fpath]
    targetpathlist = []
    targetdictpathlist = []

    for imagepath in sub1Pathlist:
        temp_list = imagepath.split('\\')#use PathLib
        filename_ele = file_key_mask.search(imagepath.split('\\')[-1]).group(0)#use PathLib
        path_ele = '\\'.join(temp_list[-3:-1])
        targetpathlist.append(('\\'.join(temp_list[-3:]))[:-4].replace(sub_sorted[0], "S$Sub_id$"))
        targetdictpathlist.append('\\'.join([path_ele, filename_ele.replace(sub_sorted[0], "S$Sub_id$")]))

    # print(f'target paths: {targetpathlist}')

    for row in zip( sub1Pathlist, targetpathlist, targetdictpathlist):
        ws2b.append(row)
    wb.save(excelfile)


    ######################  SHEET 3  ##########################################################
    # Worksheet containing the template with placeholders for user input(for all path elements)
    ws3 = wb.create_sheet('Std_Template', 4)

    # area and timepoints
    for target in targetpathlist:
        tempList = target.split('\\')

        area_group.append(area.search(tempList[-1]).group(0))
        time_group.append(time.search(tempList[-1]).group(0))

    area_sorted = sorted(set(area_group))
    time_sorted = sorted(set(time_group))
    # codes for excel file subsitutions


    # Grabbing dict values using previously generated groups as keys
    area_code_list = []
    for k in area_sorted:
        area_code_list.append(dict_code[k])

    time_code_list = []
    for k in time_sorted:
        time_code_list.append(dict_code[k])

    # writing out the template to columns and rows
    for an, ac in enumerate(area_code_list, start=2):
        ws3.cell(row=an, column=1).value = ac

    for tn, tc in enumerate(time_code_list, start=2):
        ws3.cell(row=1, column=tn).value = tc

    transfer_list = []
    for sub in sub_sorted:
        for target in Wholetargetpathlist:
            if sub in target:
                target_trimmed = target[:-4]
                transfer_list.append(target_trimmed)

    transferIter = iter(transfer_list)

    for an, ac in enumerate(area_code_list, start=2):
        for tn, tc in enumerate(time_code_list, start=2):
            ws3.cell(row=an, column=tn).value = next(transferIter)
    wb.save(excelfile)


    # Make a copy of the template for user modification
    source = wb['Std_Template']
    target = wb.copy_worksheet(source)
    target.title = "Modify_Template_here"
    wb.save(excelfile)
    filenamelist_return = [os.path.split(row)[-1] for row in filenamelist]


    return area_code_list, time_code_list, area_sorted, time_sorted, sub_sorted, filenamelist, filenamelist_return, Main_mapping_dict, transfer_list


