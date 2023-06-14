from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook

from Internal_Imports_Stable import studynumber, excelfile, Test_type
from Internal_Imports_Stable import header, header_draft, pagestyle, hypersetup
from Common_Functions_Stable import create_Latex_document
from Latex_File_Create_Stable import create_final_latex_file


def customize(area_sorted, time_sorted, sub_sorted, rownames, colnames, Main_mapping_dict, draft_flag, random_iter):
    """
    Users can customize layout as per their own preferences and study needs
    """
    # Reads the modified template
    wb = load_workbook(excelfile)
    ws4 = wb['Modify_Template_here']

    nr_col=len(colnames)#len(time_list)
    nr_row=len(rownames)#len(area_list)

    max_height=round(1.0/(1.15*nr_row), 2)
    max_width=round(1.0/(1.2*nr_col), 2)


    lax_document, new_page_alert = create_Latex_document(colnames, rownames, sub_sorted,  max_height, max_width, 
                                                         area_sorted, time_sorted, ws4, Main_mapping_dict, Type="Custom", RandomList=random_iter)


    custom_tex_file = create_final_latex_file(studynumber, header, header_draft, pagestyle, 
                                              hypersetup, lax_document, new_page_alert, colnames, Test_type, draft=draft_flag)
    
    return custom_tex_file